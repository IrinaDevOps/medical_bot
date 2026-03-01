import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO
from datetime import datetime
from services.database import patients_db
from services.health_reports import get_patient_reports

async def export_patients_to_excel() -> BytesIO:
    """Генерирует Excel файл со списком всех пациентов и их отчетами"""
    patients = await patients_db.read()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Пациенты"
    
    # Headers
    headers = [
        "ФИО", 
        "Отделение", 
        "ID Telegram", 
        "Зарегистрирован", 
        "Дата операции", 
        "Название операции", 
        "Время уведомлений", 
        "Дата удаления",
        "ID в системе",
        "Отчет (5 день)",
        "Отчет (10 день)",
        "Отчет (30 день)",
        "Другие отчеты"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Data
    row_num = 2
    for p_id, data in patients.items():
        ws.cell(row=row_num, column=1, value=data.get("full_name"))
        ws.cell(row=row_num, column=2, value=data.get("department"))
        ws.cell(row=row_num, column=3, value=data.get("user_id"))
        
        # Formatting dates
        reg_date = data.get("registration_date")
        if reg_date:
            try:
                dt = datetime.fromisoformat(reg_date)
                reg_date = dt.strftime("%d.%m.%Y %H:%M")
            except: pass
        ws.cell(row=row_num, column=4, value=reg_date)
        
        surg_date = data.get("surgery_date")
        if surg_date:
            try:
                dt = datetime.fromisoformat(surg_date)
                surg_date = dt.strftime("%d.%m.%Y")
            except: pass
        ws.cell(row=row_num, column=5, value=surg_date)
        
        ws.cell(row=row_num, column=6, value=data.get("surgery_name"))
        ws.cell(row=row_num, column=7, value=data.get("reminder_time"))
        
        del_date = data.get("auto_delete_date")
        if del_date:
             try:
                dt = datetime.fromisoformat(del_date)
                del_date = dt.strftime("%d.%m.%Y")
             except: pass
        ws.cell(row=row_num, column=8, value=del_date)
        ws.cell(row=row_num, column=9, value=p_id)
        
        # Reports
        reports_data = await get_patient_reports(p_id)
        reports_map = reports_data.get("reports", {})
        
        other_reports = []
        
        # 5, 10, 30 days mappings
        days_map = {"5": 10, "10": 11, "30": 12}
        
        for r_day, r_info in reports_map.items():
            text = r_info.get("text", "")
            date = r_info.get("submitted_at", "")
            is_urgent = r_info.get("is_urgent", False)
            
            # Format report string
            try:
                dt = datetime.fromisoformat(date)
                date_fmt = dt.strftime("%d.%m %H:%M")
            except: date_fmt = date
            
            report_str = f"[{date_fmt}] {text}"
            
            if r_day in days_map:
                col = days_map[r_day]
                ws.cell(row=row_num, column=col, value=report_str)
                ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True)
            else:
                prefix = "‼️ " if is_urgent else ""
                other_reports.append(f"{prefix}День {r_day}: {report_str}")
                
        if other_reports:
             val = "\n\n".join(other_reports)
             ws.cell(row=row_num, column=13, value=val)
             ws.cell(row=row_num, column=13).alignment = Alignment(wrap_text=True)
        
        row_num += 1
        
    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
             try:
                 val_len = len(str(cell.value))
                 if val_len > 100: val_len = 50 
                 if val_len > max_length:
                     max_length = val_len
             except: pass
        
        adjusted_width = (max_length + 2)
        if adjusted_width > 50: adjusted_width = 50 
        ws.column_dimensions[column].width = adjusted_width
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
